"""Anonymize a real placement slip into a regression-test fixture.

Per-fixture rules live in FIXTURES below. The script:
  1. If source is `.xls`, converts to `.xlsx` via Excel COM automation
     (openpyxl can only read `.xlsx`; many real broker slips are `.xls`).
  2. Loads the (possibly-converted) `.xlsx` with openpyxl (full mode).
  3. Applies string find/replace across every cell of every sheet.
  4. Applies per-cell overrides (when find/replace would mis-replace).
  5. Optionally scales numeric cells in headcount/SI columns.
  6. Saves to apps/web/tests/extraction/fixtures/<name>/slip.xlsx.
  7. Writes a per-fixture audit log of every change.

Usage:
  python scripts/anonymize-slip.py <fixture-name>
  python scripts/anonymize-slip.py --all
  python scripts/anonymize-slip.py --list

Adding a fixture: append an entry to FIXTURES, run the script, manually
verify the output in Excel (anonymization rules above in
apps/web/tests/extraction/fixtures/README.md), then commit.
"""

import argparse
import json
import os
import sys
import tempfile
from pathlib import Path
from typing import Any

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
FIXTURES_DIR = REPO_ROOT / "apps" / "web" / "tests" / "extraction" / "fixtures"

# Excel SaveAs format code for .xlsx (xlOpenXMLWorkbook).
_XL_OPEN_XML_WORKBOOK = 51


def convert_xls_to_xlsx(src: Path) -> Path:
    """Convert a legacy .xls file to .xlsx via Excel COM automation.

    Returns the path to a temp .xlsx file. Caller is responsible for cleanup
    (or just let the OS reap it on next reboot — they're tiny).

    Requires Microsoft Excel installed on Windows. Raises if Excel COM is not
    available or the file fails to open.
    """
    try:
        import pythoncom
        import win32com.client
    except ImportError as e:  # pragma: no cover - explicit error for non-Windows
        raise SystemExit(
            f"Source {src.name} is .xls and requires Excel COM to convert. "
            f"Install: pip install pywin32. Underlying: {e}"
        ) from e

    tmp_fd, tmp_path_str = tempfile.mkstemp(prefix="anon_", suffix=".xlsx")
    os.close(tmp_fd)
    tmp_path = Path(tmp_path_str)
    if tmp_path.exists():
        tmp_path.unlink()  # Excel SaveAs requires the target NOT to exist.

    pythoncom.CoInitialize()
    xl = win32com.client.Dispatch("Excel.Application")
    xl.Visible = False
    xl.DisplayAlerts = False
    try:
        wb = xl.Workbooks.Open(str(src), ReadOnly=True)
        wb.SaveAs(str(tmp_path), FileFormat=_XL_OPEN_XML_WORKBOOK)
        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()

    return tmp_path

# Per-fixture anonymization rules. Each fixture's `source` is the real slip's
# absolute path on the operator's machine; `dest` is the in-repo fixture path.
# String replacements run on every text cell (case-sensitive). Cell overrides
# run after string replacements (last-write-wins).
FIXTURES: dict[str, dict[str, Any]] = {
    "cbre-mcst-2026": {
        "source": r"C:\Users\huien\Desktop\slips\Placement Slips - CBRE MCST  (2025-2026).xlsx",
        "replace_strings": {
            # Direct identifiers
            "MCST 1216": "Test B Pte Ltd",
            # Parent group references (in "Policyholder(s) Rated Together")
            "CBRE Group": "Test Group",
            "CBRE Pte Ltd": "Test Pte Ltd",
            # Addresses (two distinct ones used by this slip)
            "6 Battery Road #32-01 Singapore 049909": "1 Test Avenue, Singapore 100001",
            "2 Tanjong Katong Road, #06-01 Paya Lebar Quarter, 437161": "2 Test Boulevard, Singapore 100002",
            # Policy numbers
            "ZZG8001400SN": "TEST/ZURICH/MCST-001",
        },
        "cell_overrides": {},
        "scale_numeric": [],
    },
    "png-2026": {
        "source": r"C:\Users\huien\Desktop\slips\Papua New Guinea - Placement Slips 2026.xlsx",
        "replace_strings": {
            # Direct identifier (appears on every product sheet at C4/C5/B21/B25/B27 etc.
            # and on the Renewal Overall Premium summary at B2).
            "Papua New Guinea High Commission": "Test C Pte Ltd",
            # Address (only one in this slip).
            "1 Marine Parade Central, #08-05 Parkway Centre Singapore 449408": "3 Test Crescent, Singapore 100003",
            # Allied World policy numbers (string, anonymized to neutral form —
            # avoid abbreviations like 'PNG' which themselves identify the source).
            "BAPASB0081902502": "TEST/AW/2026-002",
            "BWWCSB0118792502": "TEST/AW/2026-003",
        },
        "cell_overrides": {
            # Tokio Marine Life policy number is stored as an INTEGER (50011652) on
            # GTL/GHS/GCGP/GCSP/GD!C11. The string-replace pass only touches text cells,
            # so we override these explicitly.
            "GTL": {"C11": "TEST/TM/2026-001"},
            "GHS": {"C11": "TEST/TM/2026-001"},
            "GCGP": {"C11": "TEST/TM/2026-001"},
            "GCSP": {"C11": "TEST/TM/2026-001"},
            "GD": {"C11": "TEST/TM/2026-001"},
            # Round the WICA estimated annual earnings to clean numbers — exact salaries
            # are mildly identifying. Ratio preserved (Class 1 ~ 60000, Class 2 ~ 55000).
            # H32/H33 are the basis-of-cover declared wages; F37/F38 are formula refs
            # `=H32`/`=H33` in the rate-calculation table that get FLATTENED to their
            # cached value before this override runs, so we override them too.
            # H37/H38 are `=F37*G37` etc. — recompute the products manually.
            # Note: WICA Annual Premium does NOT reconcile to wages × rate (slip applies
            # a SGD250 minimum); these rounded values preserve that mismatch, see notes.md.
            "WICI": {
                "H32": 60000,
                "H33": 55000,
                "F37": 60000,
                "F38": 55000,
                "H37": 25.2,    # 60000 * 0.00042
                "H38": 137.5,   # 55000 * 0.0025
            },
        },
        "scale_numeric": [],
    },
    "vdl-2026": {
        "source": r"C:\Users\huien\Desktop\slips\VDL - Placement Slips 2026 (as at 13 Apr 2026).xls",
        "replace_strings": {
            # Three legal entities in the slip's "Insured" field. Replacement order
            # matters: longer strings FIRST so substring overlap doesn't fire early.
            # Note the source has a typo "Tecnologies" on entity 2 — replace it
            # explicitly so the output doesn't carry the misspelling.
            "VDL Enabling Technologies Group (Singapore) Pte Ltd": "Test D Pte Ltd",
            "VDL Enabling Tecnologies Group of Suzhou Ltd": "Test D China Ltd",
            "VDL Enabling Technologies Group USA LLC": "Test D Americas LLC",
            # Alternate abbreviations of the entity name appear in extension notes.
            "VDL Enabling Tech. Group of Suzhou Ltd": "Test D China Ltd",
            # Three address locations.
            "Main Location 1: 259 Jalan Ahmad Ibrahim Singapore 629148": "Main Location 1: 4 Test Lane, Singapore 100004",
            "Location 2: 16C Tuas Avenue 1 #03-41 Jtc Space @ Tuas Singapore 639535": "Location 2: 5 Test Industrial Park #01-01, Singapore 100005",
            "Location 3: Mapletree Logistics, 5A Joo Koon Circle, #04-04, Singapore 628535": "Location 3: 6 Test Logistics Hub, Singapore 100006",
            # Business description — generalize the SSIC category. The slip says
            # "Manufacturing of high tech equipment for semi-conductors & related industry";
            # we replace with a broader "electronic components" descriptor so the
            # combination of headcount + industry doesn't pinpoint the firm.
            "Manufacturing of high tech equipment for semi-conductors & related industry": "Manufacturing of electronic components",
            # Berkshire policy numbers (string).
            "47-AAH-307299": "TEST/BRK/2026-002",
            "47-ACA-306944-08": "TEST/BRK/2026-003",
            # ── Extension-notes leaks ───────────────────────────────────────────
            # "Additional Arrangements" sections at the bottom of every product
            # sheet contain free-text extensions. These leak: real employee names
            # (PII), the Dutch parent location, the Suzhou subsidiary location,
            # and a long list of third-party staffing companies. Replace each.
            #
            # Real employee names (PII — must be removed).
            "Teoh Kok Chuan": "[Employee A]",
            "Tey Xin Han": "[Employee B]",
            # Foreign-affiliate location identifiers.
            "VDL ETG Eindhoven in Netherlands": "Test D affiliate in Country C",
            "another VDL entity in the Netherlands": "another Test D entity in Country C",
            # Suzhou must run AFTER the longer "VDL Enabling Tech. Group of Suzhou Ltd"
            # replacement above so it doesn't truncate that match prematurely.
            "Suzhou China": "Country B",
            # Third-party staffing companies. Real SG firms — naming them in
            # the fixture exposes their commercial relationships with this client.
            # Note the double-space in "PersolKelly  Singapore" is preserved
            # verbatim from the source slip.
            "JOBEE PTE LTD": "Test Staffing Partner 1 Pte Ltd",
            "Recruit Express Pte Ltd": "Test Staffing Partner 2 Pte Ltd",
            "CAPITA Pte Ltd": "Test Staffing Partner 3 Pte Ltd",
            "ACExcellent Consulting Pte Ltd": "Test Staffing Partner 4 Pte Ltd",
            "Adecco Personnel Pte Ltd": "Test Staffing Partner 5 Pte Ltd",
            "Achieve Career Consultant Pte Ltd": "Test Staffing Partner 6 Pte Ltd",
            "Corestaff Pte Ltd": "Test Staffing Partner 7 Pte Ltd",
            "Rapid Recruitment Asia Pte Ltd": "Test Staffing Partner 8 Pte Ltd",
            "MCI OUTSOURCING PTE LTD": "Test Staffing Partner 9 Pte Ltd",
            "CVista HR Consulting Pte Ltd": "Test Staffing Partner 10 Pte Ltd",
            "PersolKelly  Singapore Pte Ltd": "Test Staffing Partner 11 Pte Ltd",
            "Talentvis Singapore Pte Ltd": "Test Staffing Partner 12 Pte Ltd",
            # Catch-all for any remaining bare "VDL" reference (e.g. "with VDL.",
            # "for VDL"). Run LAST so longer phrases match first.
            "VDL": "Test D",
        },
        "cell_overrides": {
            # Tokio Marine Life policy number stored as INTEGER (50011843) on every
            # Tokio Marine sheet (GTL, GHS-Locals, GHS-Secondees, GHS-Dependants,
            # GMM, GCGP, GCSP). String-replace pass skips numerics.
            "GTL ": {"C11": "TEST/TM/2026-001"},
            "GHS - Locals": {"C11": "TEST/TM/2026-001"},
            "GHS - Secondees": {"C11": "TEST/TM/2026-001"},
            "GHS - Dependants": {"C11": "TEST/TM/2026-001"},
            "GMM": {"C11": "TEST/TM/2026-001"},
            "GCGP": {"C11": "TEST/TM/2026-001"},
            "GCSP": {"C11": "TEST/TM/2026-001"},
        },
        "scale_numeric": [],
    },
    # Other fixtures appended in subsequent sessions:
    # 'hartree-2026': { ... },
    # 'stmicro-2026': { ... },
}


def anonymize_fixture(name: str) -> dict[str, Any]:
    if name not in FIXTURES:
        raise SystemExit(
            f"Unknown fixture {name!r}. Known: {sorted(FIXTURES)}. "
            f"Add an entry to FIXTURES in {Path(__file__).name}."
        )
    cfg = FIXTURES[name]
    src = Path(cfg["source"])
    dest_dir = FIXTURES_DIR / name
    dest = dest_dir / "slip.xlsx"

    if not src.exists():
        raise SystemExit(f"Source slip not found: {src}")

    print(f"[anonymize] {name}")
    print(f"  src:  {src}")
    print(f"  dest: {dest}")

    # If source is legacy .xls, convert to .xlsx via Excel COM first.
    # The fixture is always saved as .xlsx regardless of source format.
    converted_temp: Path | None = None
    load_src = src
    if src.suffix.lower() == ".xls":
        print("  note: .xls source - converting to .xlsx via Excel COM...")
        converted_temp = convert_xls_to_xlsx(src)
        load_src = converted_temp
        print(f"  temp: {converted_temp}")

    # Two-load pattern:
    #   wb_values (data_only=True)  → resolve formulas to cached values.
    #     Used to (a) detect identifiers hidden behind `=C4`-style refs,
    #     and (b) snapshot every cell's CACHED VALUE so we can flatten
    #     formulas to literals on save.
    #   wb (data_only=False)        → writable workbook.
    #
    # Why flatten formulas to literals: openpyxl does NOT recalculate
    # formulas on save and does NOT preserve cached values either. After
    # `wb.save(...)`, every formula cell's cached value is wiped — readers
    # using `data_only=True` (which the heuristic parser does) see None
    # until Excel re-opens-and-saves the file. For test fixtures, we just
    # don't need formulas at all; literals are sufficient and readable.
    wb_values = openpyxl.load_workbook(load_src, data_only=True)

    # Snapshot of (sheet, coord) → cached value, for every populated cell.
    cached_values: dict[tuple[str, str], Any] = {}
    cells_to_change: list[tuple[str, str, str, str]] = []  # for string replacements
    replace_strings = cfg.get("replace_strings", {})
    for sheet_name in wb_values.sheetnames:
        sh = wb_values[sheet_name]
        for row in sh.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cached_values[(sheet_name, cell.coordinate)] = cell.value
                if not isinstance(cell.value, str):
                    continue
                original = cell.value
                new_value = original
                for find, replace in replace_strings.items():
                    if find in new_value:
                        new_value = new_value.replace(find, replace)
                if new_value != original:
                    cells_to_change.append((sheet_name, cell.coordinate, new_value, original))
    wb_values.close()

    wb = openpyxl.load_workbook(load_src)

    changes: list[dict[str, Any]] = []

    # 1. Flatten every formula cell to its cached literal value so save
    # doesn't strip the cache. This runs BEFORE string replacements so
    # the replacement step writes to a now-literal cell, not a formula.
    formulas_flattened = 0
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        for row in sh.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    cached = cached_values.get((sheet_name, cell.coordinate))
                    if cached is not None:
                        cell.value = cached
                        formulas_flattened += 1
    if formulas_flattened > 0:
        changes.append(
            {
                "sheet": "*",
                "cell": "*",
                "kind": "flatten_formulas",
                "before": f"{formulas_flattened} formula cells",
                "after": "literal cached values",
            }
        )

    # 2. Apply string replacements (overwriting whatever's there now).
    for sheet_name, coord, new_value, original in cells_to_change:
        sh = wb[sheet_name]
        sh[coord] = new_value
        changes.append(
            {
                "sheet": sheet_name,
                "cell": coord,
                "kind": "string_replace",
                "before": original,
                "after": new_value,
            }
        )

    # 2. Per-cell overrides (run after string replacements).
    for sheet_name, cell_overrides in cfg.get("cell_overrides", {}).items():
        if sheet_name not in wb.sheetnames:
            print(f"  [warn] cell_override targets missing sheet {sheet_name!r}; skipped")
            continue
        sh = wb[sheet_name]
        for cell_addr, new_value in cell_overrides.items():
            old = sh[cell_addr].value
            if old != new_value:
                changes.append(
                    {
                        "sheet": sheet_name,
                        "cell": cell_addr,
                        "kind": "cell_override",
                        "before": old,
                        "after": new_value,
                    }
                )
                sh[cell_addr] = new_value

    # 3. Numeric scaling (e.g. headcount in a basis-of-cover table).
    for rule in cfg.get("scale_numeric", []):
        sheet_name = rule["sheet"]
        if sheet_name not in wb.sheetnames:
            continue
        sh = wb[sheet_name]
        factor = rule["factor"]
        for cell_addr in rule["cells"]:
            old = sh[cell_addr].value
            if isinstance(old, (int, float)):
                new_value = round(old * factor, 4)
                # Preserve int when factor produces an integer
                if isinstance(old, int) and float(new_value).is_integer():
                    new_value = int(new_value)
                changes.append(
                    {
                        "sheet": sheet_name,
                        "cell": cell_addr,
                        "kind": "scale_numeric",
                        "before": old,
                        "after": new_value,
                    }
                )
                sh[cell_addr] = new_value

    # 4. Save.
    dest_dir.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    wb.close()

    # 5. Write audit log to fixture directory.
    audit_path = dest_dir / "_anonymization-audit.json"
    audit_path.write_text(
        json.dumps(
            {
                "fixture": name,
                "source_filename": src.name,
                "dest_filename": dest.name,
                "rules_applied": {
                    "string_replacements": len(replace_strings),
                    "cell_overrides": sum(len(v) for v in cfg.get("cell_overrides", {}).values()),
                    "numeric_scalings": sum(len(r["cells"]) for r in cfg.get("scale_numeric", [])),
                },
                "total_changes": len(changes),
                "changes": changes,
            },
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    # Clean up the temp .xlsx if we converted from .xls.
    if converted_temp is not None and converted_temp.exists():
        try:
            converted_temp.unlink()
        except OSError:
            pass  # leave it; tempdir gets reaped

    print(f"  changes: {len(changes)}")
    print(f"  audit:   {audit_path.relative_to(REPO_ROOT)}")
    return {"name": name, "changes": len(changes), "dest": str(dest)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    g = parser.add_mutually_exclusive_group(required=True)
    g.add_argument("fixture", nargs="?", help="Fixture name (e.g. cbre-mcst-2026)")
    g.add_argument("--all", action="store_true", help="Anonymize all fixtures with rules defined")
    g.add_argument("--list", action="store_true", help="List known fixtures")
    args = parser.parse_args()

    if args.list:
        for name in sorted(FIXTURES):
            print(name)
        return 0

    targets = sorted(FIXTURES) if args.all else [args.fixture]
    for name in targets:
        anonymize_fixture(name)
    return 0


if __name__ == "__main__":
    sys.exit(main())
